"""
report.py -- Report Formatter & File Saver for ClientIQ
Generates professional Markdown reports from agent research results.
Supports both modes:
  - Company BI Report  (build_report)
  - Startup Research Report (build_startup_report)

Additionally exports reports to styled Excel (.xlsx) format.
"""

import os
from datetime import datetime
from rich.console import Console
from rich.panel import Panel
from rich.markdown import Markdown

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

console = Console()


# ═══════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════

def generate_excel(results: dict, filename: str = None) -> str:
    """
    Core Excel generator.
    results = {
        "company_info": {"Company": ..., "Website": ..., "Industry": ..., "Location": ...},
        "parameters":   {"Section Title": "Research Finding", ...}
    }
    Returns saved filepath.
    """
    if not _EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    company_name = results["company_info"].get("Company", "Unknown")

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"ClientIQ_{company_name.replace(' ', '_')}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Business Intelligence Report"

    # ── Colours ──────────────────────────────────────────────
    dark_blue   = "1B3A6B"
    accent_blue = "2E86AB"
    light_blue  = "D6EAF8"
    white       = "FFFFFF"
    light_grey  = "F2F2F2"

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80

    # ── Helper: thin border ───────────────────────────────────
    thin = Border(
        left   = Side(style="thin", color="CCCCCC"),
        right  = Side(style="thin", color="CCCCCC"),
        top    = Side(style="thin", color="CCCCCC"),
        bottom = Side(style="thin", color="CCCCCC"),
    )

    # ── TITLE BLOCK ───────────────────────────────────────────
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value     = "🔍 ClientIQ — Business Intelligence Report"
    title_cell.font      = Font(name="Calibri", size=18, bold=True, color=white)
    title_cell.fill      = PatternFill("solid", fgColor=dark_blue)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    # ── COMPANY INFO BLOCK ────────────────────────────────────
    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value     = "Company Overview"
    sub.font      = Font(name="Calibri", size=12, bold=True, color=white)
    sub.fill      = PatternFill("solid", fgColor=accent_blue)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 25

    company_fields = [
        ("Company Name", results["company_info"].get("Company",  "N/A")),
        ("Website",      results["company_info"].get("Website",  "N/A")),
        ("Industry",     results["company_info"].get("Industry", "N/A")),
        ("Location",     results["company_info"].get("Location", "N/A")),
        ("Report Date",  datetime.now().strftime("%B %d, %Y")),
    ]

    for i, (label, value) in enumerate(company_fields):
        row  = 3 + i
        fill = PatternFill("solid", fgColor=light_blue if i % 2 == 0 else white)

        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = Font(name="Calibri", size=11, bold=True, color=dark_blue)
        lc.fill      = fill
        lc.alignment = Alignment(vertical="center", indent=1)
        lc.border    = thin

        vc = ws.cell(row=row, column=2, value=value)
        vc.font      = Font(name="Calibri", size=11)
        vc.fill      = fill
        vc.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        vc.border    = thin
        ws.row_dimensions[row].height = 22

    # ── SPACER ────────────────────────────────────────────────
    spacer_row = 3 + len(company_fields)
    ws.row_dimensions[spacer_row].height = 15

    # ── PARAMETERS HEADER ─────────────────────────────────────
    header_row = spacer_row + 1
    ws.merge_cells(f"A{header_row}:B{header_row}")
    hr = ws[f"A{header_row}"]
    hr.value     = "Business Intelligence Parameters"
    hr.font      = Font(name="Calibri", size=12, bold=True, color=white)
    hr.fill      = PatternFill("solid", fgColor=accent_blue)
    hr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[header_row].height = 25

    # Column label row
    col_label_row = header_row + 1
    for col, label in enumerate(["Parameter", "Research Finding"], 1):
        c = ws.cell(row=col_label_row, column=col, value=label)
        c.font      = Font(name="Calibri", size=11, bold=True, color=white)
        c.fill      = PatternFill("solid", fgColor=dark_blue)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin
    ws.row_dimensions[col_label_row].height = 22

    # ── PARAMETER ROWS ────────────────────────────────────────
    data_start = col_label_row + 1
    for i, (param, answer) in enumerate(results["parameters"].items()):
        row  = data_start + i
        fill = PatternFill("solid", fgColor=light_grey if i % 2 == 0 else white)

        pc = ws.cell(row=row, column=1, value=param)
        pc.font      = Font(name="Calibri", size=10, bold=True, color=dark_blue)
        pc.fill      = fill
        pc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        pc.border    = thin

        ac = ws.cell(row=row, column=2, value=answer)
        ac.font      = Font(name="Calibri", size=10)
        ac.fill      = fill
        ac.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ac.border    = thin
        ws.row_dimensions[row].height = 60

    # ── FOOTER ────────────────────────────────────────────────
    footer_row = data_start + len(results["parameters"]) + 1
    ws.merge_cells(f"A{footer_row}:B{footer_row}")
    fc = ws[f"A{footer_row}"]
    fc.value     = f"Generated by ClientIQ  •  {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    fc.font      = Font(name="Calibri", size=9, italic=True, color="888888")
    fc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[footer_row].height = 20

    wb.save(filename)
    return filename


# ── Adapters ─────────────────────────────────────────────────

def save_report_excel(company_info: dict, sections: dict, output_dir: str = ".") -> str:
    """Save BI report as styled Excel. Returns filepath."""
    safe = "".join(c if c.isalnum() or c in " _-" else "_"
                   for c in company_info.get("company_name", "Report")).strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath  = os.path.join(output_dir, f"BI_Report_{safe}_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)

    results = {
        "company_info": {
            "Company":  company_info.get("company_name", "Unknown"),
            "Website":  company_info.get("website_url",  "N/A"),
            "Industry": company_info.get("industry",     "N/A"),
            "Location": company_info.get("other_details", "N/A") or "N/A",
        },
        "parameters": {
            "1. Company Overview":          sections.get("company_overview",         "Not available."),
            "2. Brand Analysis":            sections.get("brand_analysis",           "Not available."),
            "3. Website & SEO Analysis":    sections.get("website_analysis",         "Not available."),
            "4. Competitor Analysis":       sections.get("competitor_analysis",      "Not available."),
            "5. SEO & Digital Presence":    sections.get("seo_and_digital_presence", "Not available."),
            "6. Target Audience":           sections.get("target_audience",          "Not available."),
            "7. SWOT Analysis":             sections.get("swot_analysis",            "Not available."),
            "8. Market Opportunities":      sections.get("market_opportunities",     "Not available."),
            "9. Strategic Recommendations": sections.get("strategic_recommendations","Not available."),
            "10. Final Assessment":         sections.get("final_assessment",         "Not available."),
        },
    }
    return generate_excel(results, filepath)


def save_startup_report_excel(idea_info: dict, sections: dict, output_dir: str = ".") -> str:
    """Save Startup Research report as styled Excel. Returns filepath."""
    safe = "".join(c if c.isalnum() or c in " _-" else "_"
                   for c in idea_info.get("business_idea", "startup")).strip().replace(" ", "_")[:40]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath  = os.path.join(output_dir, f"Startup_Report_{safe}_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)

    results = {
        "company_info": {
            "Company":  idea_info.get("business_idea",  "Startup"),
            "Website":  "N/A",
            "Industry": idea_info.get("industry",       "N/A"),
            "Location": idea_info.get("geography",      "India"),
        },
        "parameters": {
            "1. Market Landscape":             sections.get("market_landscape",        "Not available."),
            "2. Competitor Benchmarking":      sections.get("competitor_benchmarking", "Not available."),
            "3. Target Customer Profile":      sections.get("customer_profile",        "Not available."),
            "4. Business Model & Revenue":     sections.get("business_model_options",  "Not available."),
            "5. Tech Stack & Tools":           sections.get("tech_and_tools",          "Not available."),
            "6. Marketing Channels":           sections.get("marketing_channels",      "Not available."),
            "7. Risk Analysis":                sections.get("risk_analysis",           "Not available."),
            "8. SWOT Analysis":                sections.get("startup_swot",            "Not available."),
            "9. Go-To-Market Strategy":        sections.get("go_to_market_strategy",   "Not available."),
            "10. Startup Viability Assessment":sections.get("viability_assessment",    "Not available."),
        },
    }
    return generate_excel(results, filepath)


def build_report(company_info: dict, sections: dict) -> str:
    """
    Assemble all researched sections into a structured Markdown BI report.
    Returns the full report as a string.
    """
    company_name = company_info.get("company_name", "Unknown Company")
    website_url  = company_info.get("website_url", "")
    industry     = company_info.get("industry", "")
    generated_at = datetime.now().strftime("%B %d, %Y — %H:%M")

    # ─── Header ─────────────────────────────────────────────
    report = f"""# BUSINESS INTELLIGENCE REPORT
**Company:** {company_name}
**Industry:** {industry}
**Website:** {website_url}
**Generated by:** ClientIQ — AI Business Intelligence Agent
**Date:** {generated_at}

---

> *This report was autonomously generated using live internet research. 
> All data is sourced from publicly available information.*

---

## 1. Company Overview

{sections.get("company_overview", "Not available.")}

---

## 2. Brand Analysis

{sections.get("brand_analysis", "Not available.")}

---

## 3. Website Analysis

{sections.get("website_analysis", "Not available.")}

---

## 4. Competitor Analysis

{sections.get("competitor_analysis", "Not available.")}

---

## 5. SEO & Digital Presence

{sections.get("seo_and_digital_presence", "Not available.")}

---

## 6. Target Audience

{sections.get("target_audience", "Not available.")}

---

## 7. SWOT Analysis

{sections.get("swot_analysis", "Not available.")}

---

## 8. Market Opportunities

{sections.get("market_opportunities", "Not available.")}

---

## 9. Strategic Recommendations

{sections.get("strategic_recommendations", "Not available.")}

---

## 10. Final Business Assessment

{sections.get("final_assessment", "Not available.")}

---

*— End of Report — Generated by ClientIQ*
"""

    return report


def save_report(report: str, company_name: str, output_dir: str = ".") -> str:
    """
    Save the report as a Markdown file.
    Returns the file path.
    """
    # Sanitize company name for filename
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_"
                        for c in company_name).strip().replace(" ", "_")
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"BI_Report_{safe_name}_{timestamp}.md"
    filepath   = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(report)

    return filepath


def display_report_summary(company_info: dict, sections: dict):
    """
    Print a beautiful summary of the completed report to the terminal.
    """
    company_name = company_info.get("company_name", "Unknown Company")

    console.print("\n")
    console.print(Panel(
        f"[bold green]✓ Report Complete![/]\n\n"
        f"[bold]{company_name}[/] — Business Intelligence Report\n\n"
        f"Sections researched: [cyan]{len(sections)}[/]\n"
        f"Research method: [cyan]Live internet data via ReAct agent[/]",
        title="[bold white]ClientIQ Report Summary[/]",
        border_style="green",
        padding=(1, 4)
    ))

    # Show brief snippet from each section
    section_names = {
        "company_overview":        "1. Company Overview",
        "brand_analysis":          "2. Brand Analysis",
        "website_analysis":        "3. Website Analysis",
        "competitor_analysis":     "4. Competitor Analysis",
        "seo_and_digital_presence":"5. SEO & Digital Presence",
        "target_audience":         "6. Target Audience",
        "swot_analysis":           "7. SWOT Analysis",
        "market_opportunities":    "8. Market Opportunities",
        "strategic_recommendations":"9. Strategic Recommendations",
        "final_assessment":        "10. Final Assessment",
    }

    console.print("\n[bold]Section Status:[/]")
    for key, display_name in section_names.items():
        content = sections.get(key, "")
        status  = "[green]OK[/]" if content and len(content) > 50 else "[red]--[/]"
        preview = content[:60].replace("\n", " ") + "..." if content else "Empty"
        console.print(f"  {status} {display_name}")
        console.print(f"     [dim]{preview}[/]")


# ═══════════════════════════════════════════════════════════
# STARTUP REPORT BUILDER
# ═══════════════════════════════════════════════════════════

def build_startup_report(idea_info: dict, sections: dict) -> str:
    """
    Assemble all startup research sections into a Markdown report.
    Returns the full report as a string.
    """
    business_idea  = idea_info.get("business_idea", "Unknown Business Idea")
    industry       = idea_info.get("industry", "")
    target_market  = idea_info.get("target_market", "")
    geography      = idea_info.get("geography", "")
    business_model = idea_info.get("business_model", "")
    unique_angle   = idea_info.get("unique_angle", "")
    generated_at   = datetime.now().strftime("%B %d, %Y -- %H:%M")

    report = f"""# STARTUP RESEARCH REPORT
**Business Idea:** {business_idea}
**Industry:** {industry}
**Target Market:** {target_market}
**Geography:** {geography}
**Business Model:** {business_model}
**Unique Angle:** {unique_angle}
**Generated by:** ClientIQ -- Startup Research Agent
**Date:** {generated_at}

---

> *This report was autonomously generated using live internet research.*
> *All data is sourced from publicly available information.*
> *Use this as a starting point for your business planning, not as financial advice.*

---

## 1. Market Landscape

{sections.get("market_landscape", "Not available.")}

---

## 2. Competitor Benchmarking

{sections.get("competitor_benchmarking", "Not available.")}

---

## 3. Target Customer Profile

{sections.get("customer_profile", "Not available.")}

---

## 4. Business Model & Revenue Options

{sections.get("business_model_options", "Not available.")}

---

## 5. Tech Stack & Tools Required

{sections.get("tech_and_tools", "Not available.")}

---

## 6. Marketing Channels & Customer Acquisition

{sections.get("marketing_channels", "Not available.")}

---

## 7. Risk Analysis

{sections.get("risk_analysis", "Not available.")}

---

## 8. SWOT Analysis (New Entrant Perspective)

{sections.get("startup_swot", "Not available.")}

---

## 9. Go-To-Market Strategy

{sections.get("go_to_market_strategy", "Not available.")}

---

## 10. Startup Viability Assessment

{sections.get("viability_assessment", "Not available.")}

---

*-- End of Startup Research Report -- Generated by ClientIQ*
"""
    return report


def save_startup_report(report: str, business_idea: str, output_dir: str = ".") -> str:
    """Save the startup report as a Markdown file. Returns the file path."""
    safe_name = "".join(
        c if c.isalnum() or c in " _-" else "_" for c in business_idea
    ).strip().replace(" ", "_")[:40]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"Startup_Report_{safe_name}_{timestamp}.md"
    filepath  = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(report)
    return filepath


def display_startup_summary(idea_info: dict, sections: dict):
    """Print terminal summary for startup report."""
    business_idea = idea_info.get("business_idea", "Your Business Idea")

    console.print("\n")
    console.print(Panel(
        f"[bold green]Startup Research Complete![/]\n\n"
        f"[bold]{business_idea}[/]\n\n"
        f"Sections researched: [cyan]{len(sections)}[/]\n"
        f"Research method: [cyan]Live internet data via ReAct agent[/]",
        title="[bold white]ClientIQ -- Startup Report Summary[/]",
        border_style="yellow",
        padding=(1, 4)
    ))

    section_names = {
        "market_landscape":      "1. Market Landscape",
        "competitor_benchmarking":"2. Competitor Benchmarking",
        "customer_profile":      "3. Target Customer Profile",
        "business_model_options":"4. Business Model & Revenue",
        "tech_and_tools":        "5. Tech Stack & Tools",
        "marketing_channels":    "6. Marketing Channels",
        "risk_analysis":         "7. Risk Analysis",
        "startup_swot":          "8. SWOT Analysis",
        "go_to_market_strategy": "9. Go-To-Market Strategy",
        "viability_assessment":  "10. Viability Assessment",
    }

    console.print("\n[bold]Section Status:[/]")
    for key, display_name in section_names.items():
        content = sections.get(key, "")
        status  = "[green]OK[/]" if content and len(content) > 50 else "[red]--[/]"
        preview = content[:60].replace("\n", " ") + "..." if content else "Empty"
        console.print(f"  {status} {display_name}")
        console.print(f"     [dim]{preview}[/]")

